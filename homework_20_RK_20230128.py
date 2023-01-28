# Прочитать сохранённый csv-файл из задания №19 и сохранить данные в excel-файл,
# кроме возраста – столбец с этими данными не нужен.
# К заданию прикреплён пример как должно выглядеть содержания итогового файла.
import csv
import openpyxl

with open('witcher.csv', 'r', encoding='utf-8') as new_file:
    loaded_data = csv.reader(new_file, delimiter = ',')
    wb = openpyxl.Workbook()
    sheet = wb.active
    column_headers = ['', 'person1', 'person2', 'person3', 'person4', 'person5', 'person6', 'person7']
    sheet.append(column_headers)
    for col_index, row in enumerate(loaded_data):
        for row_index, value in enumerate(row):
            cell = sheet.cell(column=col_index+1, row=row_index+2)
            cell.value = value
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == 'Age':
            sheet.delete_rows(row)
    wb.save('witcher.xlsx')


